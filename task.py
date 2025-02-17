from openpyxl import load_workbook
import json

# Load the Excel file
file_path = 'Menu-16-28th-Feb.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active  # Access the active sheet

# Initialize a dictionary to store menu data
menu_data = {}

# Extract dates from the second row (row 2)
dates = [cell.value for cell in sheet[2][:15]]  # Upto 15th column (inclusive)

# Define meal types and their corresponding row ranges
meal_types = {
    "Breakfast": (4, 12),  # Rows 3 to 12 (inclusive)
    "Lunch": (15, 22),     # Rows 14 to 23 (inclusive)
    "Dinner": (25, 34)     # Rows 25 to 34 (inclusive)
}

# print(meal_types.items(),'\n')
# print(dates,'\n')

# Iterate over each date (column) and extract menu items
for col_index, date in enumerate(dates, start=1):  # Start from column A (index=1)
    day_menu = {"Breakfast": [], "Lunch": [], "Dinner": []}
    
    for meal_type, (start_row, end_row) in meal_types.items():
        items = []
        for row in range(start_row, end_row + 1):  # Iterate through rows for each meal type
            cell_value = sheet.cell(row=row, column=col_index).value
            if cell_value and "****" not in str(cell_value):  # Exclude empty cells and invalid rows
                items.append(cell_value)
        day_menu[meal_type] = items

        # print(meal_type, items,start_row, end_row,sep='\n')
        # print('\n\n')
    
    menu_data[date.strftime('%Y-%m-%d')] = day_menu  # Format date as string

# print(menu_data)

# Save menu data to a JSON file
output_file = 'mess_menu.json'
with open(output_file, 'w') as json_file:
    json.dump(menu_data, json_file, indent=4)

print(f"Menu data saved to {output_file}")
